from __future__ import annotations

import json
from pathlib import Path

import tempfile
import os

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Response

from src.core.config import Settings
from src.core.dependencies import get_settings_dep
from src.services.igg_calculator import (
    calcular_igg_por_km,
    montar_estacoes_por_km,
)
from src.services.planilha_import import importar_planilha

router = APIRouter(prefix="/api/relatorios", tags=["relatorios"])


@router.get("/viagens")
async def listar_viagens(settings: Settings = Depends(get_settings_dep)):
    base = settings.dados_dir
    if not base.is_dir():
        return []

    from src.services.igg_calculator import classificar_igg

    resultados = []
    for pasta in sorted(base.iterdir()):
        if not pasta.is_dir():
            continue
        analise_path = pasta / "analise_completa.json"
        config_path = pasta / "viagem_config.json"
        if not analise_path.is_file() or not config_path.is_file():
            continue

        try:
            config = json.loads(config_path.read_text(encoding="utf-8"))
            analise = json.loads(analise_path.read_text(encoding="utf-8"))
        except Exception:
            continue

        params_path = pasta / "parametros_adicionais.json"
        parametros: dict = {}
        if params_path.is_file():
            try:
                parametros = json.loads(params_path.read_text(encoding="utf-8"))
            except Exception:
                parametros = {}

        total_imagens = len(analise)
        igg_por_km: list[dict] = []

        agrupado = montar_estacoes_por_km(analise, pasta.name, pasta)
        for km_key, data in sorted(agrupado.items(), key=lambda x: int(x[0])):
            params_km = parametros.get(km_key, {})
            r = calcular_igg_por_km(int(km_key), data["estacoes"], params_km)
            igg_por_km.append({
                "km": r["km"],
                "igg": r["igg"],
                "conceito": r["conceito"],
            })

        igg_medio = round(sum(i["igg"] for i in igg_por_km) / len(igg_por_km), 2) if igg_por_km else 0.0
        conceitos_prioridade = ["Péssimo", "Ruim", "Regular", "Bom", "Ótimo"]
        conceito_medio = min(
            (c for c in conceitos_prioridade if c in [i["conceito"] for i in igg_por_km]),
            key=lambda c: conceitos_prioridade.index(c),
        ) if igg_por_km else ""

        resultados.append({
            "viagem": pasta.name,
            "config": config,
            "igg_por_km": igg_por_km,
            "igg_medio": igg_medio,
            "conceito_medio": conceito_medio,
            "total_imagens": total_imagens,
            "total_kms": len(igg_por_km),
        })

    return resultados


@router.post("/retigrafico")
async def gerar_retigrafico():
    return {"status": "pending", "arquivo": ""}


@router.post("/antt")
async def exportar_antt():
    return {"status": "pending", "arquivo": ""}


@router.post("/igg/{viagem_nome}")
async def calcular_igg(
    viagem_nome: str,
    body: dict,
    settings: Settings = Depends(get_settings_dep),
):
    base = settings.dados_dir / viagem_nome
    analise_path = base / "analise_completa.json"
    params_path = base / "parametros_adicionais.json"

    if not analise_path.is_file():
        raise HTTPException(404, "Análise não encontrada para esta viagem")

    try:
        analise: dict = json.loads(analise_path.read_text(encoding="utf-8"))
    except Exception as e:
        raise HTTPException(500, f"Erro ao ler análise: {e}")

    parametros: dict = {}
    if params_path.is_file():
        try:
            parametros = json.loads(params_path.read_text(encoding="utf-8"))
        except Exception:
            parametros = {}

    km_filter = body.get("km")

    agrupado = montar_estacoes_por_km(analise, viagem_nome, base)
    if not agrupado:
        return []

    resultados = []
    for km_key, data in sorted(agrupado.items(), key=lambda x: int(x[0])):
        if km_filter is not None and km_key != str(km_filter):
            continue
        params_km = parametros.get(km_key, {})
        resultado = calcular_igg_por_km(int(km_key), data["estacoes"], params_km)
        resultados.append(resultado)

    return resultados


@router.post("/importar-planilha/{viagem_nome}")
async def importar_planilha_endpoint(
    viagem_nome: str,
    file: UploadFile = File(...),
    settings: Settings = Depends(get_settings_dep),
):
    base = settings.dados_dir / viagem_nome
    if not base.is_dir():
        raise HTTPException(404, f"Viagem não encontrada: {viagem_nome}")

    # Lê config da viagem para saber tipo de pista e faixa
    config_path = base / "viagem_config.json"
    tipo_pista = "simples"
    faixa = None
    if config_path.is_file():
        try:
            config = json.loads(config_path.read_text(encoding="utf-8"))
            tipo_pista = config.get("tipo_pista", "simples")
            faixa = config.get("faixa")
        except Exception:
            pass

    params_path = base / "parametros_adicionais.json"

    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_path = Path(tmpdir) / (file.filename or "planilha.xls")
        content = await file.read()
        tmp_path.write_bytes(content)

        try:
            dados_importados = importar_planilha(tmp_path, tipo_pista, faixa)
        except ValueError as e:
            raise HTTPException(400, str(e))

    parametros_existentes: dict = {}
    if params_path.is_file():
        try:
            parametros_existentes = json.loads(params_path.read_text(encoding="utf-8"))
        except Exception:
            parametros_existentes = {}

    kms_importados = []
    for km_key, valores in dados_importados.items():
        if km_key not in parametros_existentes:
            parametros_existentes[km_key] = {}
        parametros_existentes[km_key]["TRI (mm)"] = valores["TRI (mm)"]
        parametros_existentes[km_key]["TRE (mm)"] = valores["TRE (mm)"]
        kms_importados.append(int(km_key))

    params_path.write_text(
        json.dumps(parametros_existentes, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    return {
        "status": "ok",
        "kms_importados": sorted(kms_importados),
        "total_kms": len(kms_importados),
    }


@router.post("/exportar/{viagem_nome}")
async def exportar_viagem(
    viagem_nome: str,
    settings: Settings = Depends(get_settings_dep),
):
    import io
    import csv

    base = settings.dados_dir / viagem_nome
    analise_path = base / "analise_completa.json"
    if not analise_path.is_file():
        raise HTTPException(404, "Análise não encontrada")

    try:
        analise = json.loads(analise_path.read_text("utf-8"))
    except Exception as e:
        raise HTTPException(500, f"Erro ao ler análise: {e}")

    params_path = base / "parametros_adicionais.json"
    parametros: dict = {}
    if params_path.is_file():
        try:
            parametros = json.loads(params_path.read_text("utf-8"))
        except Exception:
            parametros = {}

    agrupado = montar_estacoes_por_km(analise, viagem_nome, base)
    if not agrupado:
        return Response(content="", media_type="text/csv", headers={"Content-Disposition": "attachment; filename=export.csv"})

    resultados = []
    for km_key, data in sorted(agrupado.items(), key=lambda x: int(x[0])):
        params_km = parametros.get(km_key, {})
        r = calcular_igg_por_km(int(km_key), data["estacoes"], params_km)
        resultados.append(r)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["KM_INICIAL", "KM_FINAL", "IGG", "CONCEITO"])
    for r in resultados:
        km = r["km"]
        writer.writerow([km, km + 1, r["igg"], r["conceito"]])

    csv_content = output.getvalue()
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=igg_{viagem_nome}.csv",
            "Content-Type": "text/csv; charset=utf-8",
        },
    )


@router.post("/exportar-xlsx/{viagem_nome}")
async def exportar_viagem_xlsx(
    viagem_nome: str,
    settings: Settings = Depends(get_settings_dep),
):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    base = settings.dados_dir / viagem_nome
    analise_path = base / "analise_completa.json"
    if not analise_path.is_file():
        raise HTTPException(404, "Análise não encontrada")

    analise = json.loads(analise_path.read_text("utf-8"))
    params_path = base / "parametros_adicionais.json"
    parametros: dict = {}
    if params_path.is_file():
        try:
            parametros = json.loads(params_path.read_text("utf-8"))
        except Exception:
            parametros = {}

    agrupado = montar_estacoes_por_km(analise, viagem_nome, base)
    if not agrupado:
        raise HTTPException(404, "Nenhum KM encontrado")

    resultados = []
    for km_key, data in sorted(agrupado.items(), key=lambda x: int(x[0])):
        params_km = parametros.get(km_key, {})
        r = calcular_igg_por_km(int(km_key), data["estacoes"], params_km)
        resultados.append(r)

    wb = Workbook()
    ws = wb.active
    ws.title = f"IGG {viagem_nome}"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = ["KM INICIAL", "KM FINAL", "IGG", "CONCEITO"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    conceito_fills = {
        "Ótimo": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
        "Bom": PatternFill(start_color="CCFBF1", end_color="CCFBF1", fill_type="solid"),
        "Regular": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
        "Ruim": PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid"),
        "Péssimo": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
    }

    for i, r in enumerate(resultados, 2):
        ws.cell(row=i, column=1, value=r["km"]).border = thin_border
        ws.cell(row=i, column=2, value=r["km"] + 1).border = thin_border
        ws.cell(row=i, column=3, value=r["igg"]).border = thin_border
        cell = ws.cell(row=i, column=4, value=r["conceito"])
        cell.border = thin_border
        fill = conceito_fills.get(r["conceito"])
        if fill:
            cell.fill = fill
            cell.font = Font(bold=True)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 14

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        content=output.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=igg_{viagem_nome}.xlsx",
        },
    )
